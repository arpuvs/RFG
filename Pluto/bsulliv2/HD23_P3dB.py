# Author: Ben Sullivan
# Date: 11/06/2017

# Necessary module and instrument imports
import sys, visa, time
sys.path.append('../../Common')
sys.path.append('../../Common/FMB_USB_Python_Files')
from ADI_GPIB.AgilentN5181A import *
from ADI_GPIB.AgilentN9030A import *
from ADI_GPIB.AgilentN6705B import *
from ADI_GPIB.WatlowF4 import *
from FMB import *
from openpyxl import *

# Instrument initialization
Supply = AgilentN6705B(26)
Source = AgilentN5181A(11)
Analyzer = AgilentN9030A(18)
Oven = WatlowF4(4)
startTime = time.time()


# Main body of code, called by GUI
def HD23Main(path, supplyVlist, freqlist, vcomlist, templist, dutNumber, vcomEnable):

    def HD23():  # Main HD23 measurement function
        # Initializes analyzer
        Analyzer.__SetSpan__(10e3)
        Analyzer.__SetAverage__(50)

        # Initializes all measurement arrays
        fundamental = []
        second = []
        third = []
        Carrier = []
        SourceAmp = []
        secondNormal = []
        thirdNormal = []
        supplyI = []
        freqlistReal = []

        for freqindex in freqlist:
            # Converts fmb dict frequency to float
            freq = fmbDict[freqindex].split()
            if freq[1] == 'MHz':
                freq = float(freq[0]) * 1e6
            elif freq[1] == 'GHz':
                freq = float(freq[0]) * 1e9
            freqlistReal.append(freq)

            # Sets up all equipment for given frequency
            Filter.select_filter(freqindex)
            Source.__SetFreq__(freq)
            Analyzer.__Setfc__(freq)
            Source.__SetState__(1)
            Analyzer.__SetMarkerFreq__(1, freq)
            time.sleep(0.1)
            Analyzer.__ClearAverage__()

            # Sets initial amplitude using measured output of dut
            Analyzer.__CheckStatus__(300)                     # Waits until averaging is complete
            time.sleep(0.5)
            Analyzer.__ClearAverage__()
            Analyzer.__CheckStatus__(300)
            carrierMag = float(Analyzer.__GetMarkerAmp__(1))  # Gets initial amplitude
            while abs(carrierMag - -2) >= 0.1:                # Adjusts amplitude until it is within 0.1 dBm of desired
                sourceAmp = float(Source.__GetAmp__())
                setAmp = sourceAmp + (-2 - carrierMag)
                if setAmp > 15:
                    raise Exception('Max source amplitude exceeded')
                Source.__SetAmp__(setAmp)
                Analyzer.__ClearAverage__()
                Analyzer.__CheckStatus__(300)
                carrierMag = float(Analyzer.__GetMarkerAmp__(1))

            # Gets final amplitude
            Carrier.append(carrierMag)
            fundamental.append(carrierMag)
            SourceAmp.append(float(Source.__GetAmp__()))

            # Makes second and third harmonic measurements
            Analyzer.__Setfc__(freq*2.0)
            Analyzer.__SetMarkerFreq__(1, freq*2.0)
            Analyzer.__ClearAverage__()
            Analyzer.__CheckStatus__(300)
            second.append(float(Analyzer.__GetMarkerAmp__(1)))
            Analyzer.__Setfc__(freq*3.0)
            Analyzer.__SetMarkerFreq__(1, freq*3.0)
            Analyzer.__ClearAverage__()
            Analyzer.__CheckStatus__(300)
            third.append(float(Analyzer.__GetMarkerAmp__(1)))
            supplyI.append(float(Supply.__GetI__(1)))

        # Converts dBm measurements to dBc
        for j in range(len(second)):
            secondNormal.append(-(float(Carrier[j]) - float(second[j])))
            thirdNormal.append(-(float(Carrier[j]) - float(third[j])))

        # Appends measured data at end of specified xlsx file
        Acolumn = sheet_ranges['A']
        index = 0
        for row in range(len(Acolumn)+1, len(Acolumn) + len(freqlistReal) + 1):
            data = [dutNumber, channel, temp, supplyV, vcom, atten, freqlistReal[index], supplyI[index], Carrier[index],
                    second[index], third[index], secondNormal[index], thirdNormal[index]]
            for col in range(len(data)):
                sheet_ranges.cell(column=col+1, row=row, value=data[col])
            index = index + 1

    # Sets oven to setpoint and soaks dut for allotted time
    def setTemp(setpoint):
        Oven.__SetTemp__(setpoint)
        current = float(Oven.__GetTemp__())
        while (abs(current - setpoint) > 2):
            time.sleep(1)
            current = float(Oven.__GetTemp__())
        print '@ Temp %d' % setpoint
        soak = 300
        time.sleep(soak)
        return True

    # Sets analyzer attenuation settings (22dB was empirically chosen)
    Analyzer.__SetAutoAtten__(0)
    Analyzer.__SetAtten__(22)

    # Filter box frequency settings dictionary
    fmbDict = {1: "2.5 MHz", 2: "5 MHz", 3: "33 MHz", 4: "78 MHz",
              5: "120 MHz", 6: "225.3 MHz", 7: "350.3 MHz", 8: "500 MHz",
              9: "800.3 MHz", 10: "1 GHz", 11: "1.5 GHz", 12: "2 GHz",
              13: "2.5 GHz", 14: "3 GHz", 15: "3.5 GHz", 16: "4 GHz",
              17: "4.5 GHz", 18: "5 GHz", 19: "5.5 GHz", 20: "5.9 GHz",
              21: "Aux"}

    Filter = FMB('COM3', fmbDict)  # Initializes filter box

    # Opens specified xlsx file. If no file exists a new spreadsheet is created.
    try:
        wb = load_workbook(filename=summaryPath)
        sheet_ranges = wb['Sheet1']
    except:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = 'Sheet1'
        sheet_ranges = wb['Sheet1']
        firstline = ['DUT', 'Channel', 'Temp', 'Supply', 'Vcom', 'Attenuation', 'Frequency', 'Current', 'Carrier(dB)',
                     'Second(dB)', 'Third(dB)', 'HD2(dBc)', 'HD3(dBc)']
        col = 1
        for item in firstline:
            sheet_ranges.cell(column=col, row=1, value=firstline[col-1])
            col = col + 1


    # Sets up main supply
    Supply.__SetV__(supplyVlist[0], 1)
    Supply.__SetI__(0.25, 1)
    Supply.__Enable__(1, 1)
    if vcomEnable:
        Supply.__Enable(1, 3)

    # Main loop structure
    for temp in templist:
        if templist != [25]:
            setTemp(temp)                       # Sets DUT to temperature
        for supplyV in supplyVlist:
            Supply.__SetV__(supplyV, 1)
            for vcom in vcomlist:
                Supply.__SetV__(vcom, 3)  # Sets DUT to common mode voltage
                print 'Vcom = %g' % vcom
                for atten in attenlist:
                    pluto.Set_Amp_Atten(SPI_sel=channel, AmpAtten=atten)
                    print 'Atten = %g' % atten

                    Source.__SetState__(0)
                    print 'Aligning...'
                    Analyzer.__CheckStatus__(300)
                    Analyzer.__Align__()            # Aligns device with no input
                    Analyzer.__CheckStatus__(300)
                    print 'Done'

                    HD23()                          # Runs main HD23 measurement

    # Returns oven to ambient temp, finds execution time and saves file
    Supply.__Enable__(0, 1)
    if vcomEnable:
        Supply.__Enable(0, 3)
    if templist != [25]:
        Oven.__SetTemp__(25)
    print time.time()-startTime
    wb.save(filename=path)
    print 'Done!'


# Called if program run by itself
if __name__ == '__main__':
    # Sets all necessary variables
    path = 'C:\\Users\\#RFW_Test01\\Desktop\\5569_Data\\Hd23\\HD23.xlsx'
    # freqlist = [100e6, 250e6, 500e6, 1.0e9, 1.5e9, 2.0e9, 2.5e9, 3.0e9, 3.5e9, 4.0e9]
    freqs = [5, 6, 8, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20]
    # freqs = [12]
    # for i in range(20, 31):
    #     vcoms.append(i/10.0)
    # vcoms = [2.0, 2.5, 3.0]
    supplies = [5]
    vcoms = [2.5]
    temps = [25]
    vcomEnable = 0
    attenlist = [0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9]
    dut = '3-6'
    channel = 'B'

    instDict = InstInit()

    pluto = PlutoV1()
    pluto.connect(DUT1_Default=0x00, DUT2_Default=0x00)
    pluto.Set_Amp_Enable(SPI_sel=channel, AmpEnable=True)

    HD23Main(path, supplies, freqs, vcoms, temps, dut, vcomEnable)  # Calls main program

    pluto.Set_Amp_Enable(SPI_sel=channel, AmpEnable=False)


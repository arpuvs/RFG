# Author: Ben Sullivan
# Date: 12/19/2017

import sys, visa, time, math, cmath
sys.path.append('../../Common')

# Instrumnet imports and initialization
# from ADI_GPIB.WatlowF4 import *
# from ADI_GPIB.E3631A import *
# from ADI_GPIB.KeysightN5242A import *
from openpyxl import *
from BSTest import *
from PlutoV0 import PlutoV0

# Sets all necessary VNA parameters and adds all specified measurements
def SParam():
    def VNAinit():
        instDict['NA'].__Preset__()
        instDict['NA'].__SetSweepType__(1, 'LOG')
        instDict['NA'].__SetStartf__(1, startFreq)
        instDict['NA'].__SetStopf__(1, endFreq)
        instDict['NA'].__SetNumPoints__(1, numPoints)
        instDict['NA'].__SetAutoTime__(1, True)
        instDict['NA'].__EnableAvg__(1, False)
        instDict['NA'].__SetTopology__(1, 'BBAL')
        instDict['NA'].__SetPorts__(1, 1, 3, 2, 4)
        instDict['NA'].__EnableBal__(1)
        instDict['NA'].__RemoveTrace__(1, 1)
        trace = 1

        # Adds measurements from supplied list
        for meas in measlist:
            instDict['NA'].__AddMeas__(1, meas, 'Standard', meas.split()[0])
            instDict['NA'].__AddTrace__(1, trace, meas)
            instDict['NA'].__SetActiveTrace__(1, meas)
            instDict['NA'].__SetActiveFormat__(1, meas.split()[1])
            trace = trace + 1

        instDict['NA'].__RecallCal__('Pluto_Sparam')
        instDict['NA'].__EnableAvg__(1, True)
        instDict['NA'].__SetAvg__(1, avg)

    # Sets oven to specified temperature and soaks
    def setTemp(temperature):
            # Could add function to turn off supplies when ramping up per Greg's suggestion
            instDict['thermo'].__SetupDUTMode__(HighTemp=145, LowTemp=-70, SensorType='K', TestTime=230)
            instDict['thermo'].__SetDutDtype__(1)
            instDict['thermo'].__SetDutThermalConst__(100)
            instDict['thermo'].__EnableDUTMode__(SensorType='K')
            off = False
            instDict['thermo'].__SetTemp__(temperature)
            instDict['thermo'].__FlowON__()
            instDict['thermo'].__MoveArmDown__()
            instDict['thermo'].__EnableDUTMode__('K')
            time.sleep(5)
            measTemp = instDict['thermo'].__GetDutTemperature__()
            # if temperature > measTemp:
            #     powerDown(instDict)
            #     off = True
            while (abs(temperature - measTemp) > 5):
                time.sleep(10)
                measTemp = instDict['thermo'].__GetDutTemperature__()
                print measTemp
            print 'At temp'
            print 'Soaking...'
            time.sleep(300)
            # if off:
            #     powerUp(instDict)
            #     # chip = Linus(bridge_device='aardvark', linus_rev=2)

    # The following three functions are simple math operations taken from VNA Vee program
    def build_av(mlog):
        ans = []
        for val in range(len(mlog)):
            ans.append(mlog[val] - 10.0*math.log10(Zin_diff/Zout_diff))
        return ans


    def build_cmrr(sdd21, sdc21):
        ans = []
        for val in range(len(sdd21)):
            ans.append(sdd21[val] - sdc21[val])
        return ans


    def build_gdel(phase, freq):
        ans = []
        ans.append(0)
        for val in range(len(phase) - 1):
            ans.append(((phase[val+1] - phase[val])*math.pi/180.0)/(freq[val+1]-freq[val]))
        return ans


    # Retrieves measured data from VNA and prints to file
    def getData():
        readDict = {}   # Dictionary for results to be stored in
        instDict['NA'].__RestartAvg__(1)
        time.sleep(5)  # Necessary to let one sweep finish. Otherwise FinishAvg function does not work
        instDict['NA'].__FinishAvg__(1, 600)   # Pauses execution until VNA is finished averaging

        # Retrieve data from VNA
        for meas in measlist:
            instDict['NA'].__SetActiveTrace__(1, meas)
            ans = instDict['NA'].__GetData__(1)
            ans = ans.split(',')
            # Converts received data from unicode to numeric
            for val in range(len(ans)):
                ans[val] = float(ans[val])
            # Polar measurements return real and imaginary data that must be separated
            if meas.split()[1] == 'POL':
                readDict[meas + '1'] = []
                readDict[meas + '2'] = []
                for i in range(len(ans)):
                    if i % 2:
                        readDict[meas + '1'].append(ans[i])
                    else:
                        readDict[meas + '2'].append(ans[i])
            else:
                readDict[meas] = ans

        # Retrieves and converts swept frequency points
        freqlist = instDict['NA'].__GetFreq__(1)
        freqlist = freqlist.split(',')
        for val in range(len(freqlist)):
            freqlist[val] = float(freqlist[val])

        # Math functions from VNA Vee program
        av = build_av(readDict['SDD21 MLOG'])
        readDict['AV'] = av
        cmrr1 = build_cmrr(readDict['SDD21 MLOG'], readDict['SDC21 MLOG'])
        readDict['CMRR1'] = cmrr1
        cmrr2 = build_cmrr(readDict['SDD21 MLOG'], readDict['SCC21 MLOG'])
        readDict['CMRR2'] = cmrr2
        group_delay = build_gdel(readDict['SDD21 POL1'], freqlist)
        readDict['GDEL'] = group_delay
        s12_v = build_av(readDict['SDD12 MLOG'])
        readDict['S12_V'] = s12_v
        # zyIn = build_zy(sdd11_mlog)
        # zyOut = build_zy(sdd22_mlog)

        # Writes all data to file
        fh.write('Frequency,')
        fh.write(str(freqlist).strip('[]'))
        fh.write('\n')
        for meas in measlist:
            # Polar measurements need additional processing due to complex form
            if meas.split()[1] == 'POL':
                fh.write(meas + '1')
                fh.write(',')
                fh.write(str(readDict[meas + '1']).strip('[]'))
                fh.write('\n')
                fh.write(meas + '2')
                fh.write(',')
                fh.write(str(readDict[meas + '2']).strip('[]'))
                fh.write('\n')
            # All other measurements can be printed as is
            else:
                fh.write(meas)
                fh.write(',')
                fh.write(str(readDict[meas]).strip('[]'))
                fh.write('\n')
        fh.write('AV, ')
        fh.write(str(av).strip('[]'))
        fh.write('\n')
        fh.write('CMRR1, ')
        fh.write(str(cmrr1).strip('[]'))
        fh.write('\n')
        fh.write('CMRR2, ')
        fh.write(str(cmrr2).strip('[]'))
        fh.write('\n')
        fh.write('Group Delay,')
        fh.write(str(group_delay).strip('[]'))
        fh.write('\n')
        fh.write('S12_V,')
        fh.write(str(s12_v).strip('[]'))
        fh.write('\n')
        summary(freqlist, readDict)

    # Prints first line of file
    def header():
        test = 'P1dB'
        equipment = 'N5242A PNA-X BAL0026'
        header = (dut, date, test, equipment)
        header = str(header).strip('()')
        fh.write(header)
        fh.write('\n')

    def summary(freqlist, readDict):

        Acolumn = sheet_ranges['A']
        index = 0
        for row in range(len(Acolumn)+1, len(Acolumn) + 50):
            data = [dut, temp, supply, vcom, Pout, pwrmode, gain, freqlist[index], readDict['SCC21 MLOG'][index],
                    readDict['SDC21 MLOG'][index], readDict['SDD11 POL1'][index], readDict['SDD11 POL2'][index],
                    readDict['SDD12 MLOG'][index], readDict['SDD21 GDEL'][index], readDict['SDD21 MLOG'][index],
                    readDict['SDD21 POL1'][index], readDict['SDD21 POL2'][index], readDict['SDD11 MLOG'][index],
                    readDict['SDD22 MLOG'][index], readDict['AV'][index], readDict['CMRR1'][index],
                    readDict['CMRR2'][index], readDict['GDEL'][index], readDict['S12_V'][index]]
            for col in range(len(data)):
                sheet_ranges.cell(column=col+1, row=row, value=data[col])
            index = index + int(numPoints/50)

        index = len(freqlist)-1
        data = [dut, temp, supply, vcom, Pout, pwrmode, gain, freqlist[index], readDict['SCC21 MLOG'][index],
                readDict['SDC21 MLOG'][index], readDict['SDD11 POL1'][index], readDict['SDD11 POL2'][index],
                readDict['SDD12 MLOG'][index], readDict['SDD21 GDEL'][index], readDict['SDD21 MLOG'][index],
                readDict['SDD21 POL1'][index], readDict['SDD21 POL2'][index], readDict['SDD11 MLOG'][index],
                readDict['SDD22 MLOG'][index], readDict['AV'][index], readDict['CMRR1'][index],
                readDict['CMRR2'][index], readDict['GDEL'][index], readDict['S12_V'][index]]
        for col in range(len(data)):
            sheet_ranges.cell(column=col + 1, row=row+1, value=data[col])


    startTime = time.time()

    date = time.ctime(time.time())
    date = date.replace(':', '-')

    fh = open(path + dut + ' SParam_' + date + '.csv', 'w')    # Creates csv file

    try:
        wb = load_workbook(filename=summaryPath)
        sheet_ranges = wb['Sheet1']
    except:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = 'Sheet1'
        sheet_ranges = wb['Sheet1']
        firstline = ['DUT', 'Temp', 'Supply', 'Vcom', 'Pout', 'PowerMode', 'Gain', 'Frequency', 'SCC21 MLOG', 'SDC21 MLOG',
                     'SDD11 POL1', 'SDD11 POL2', 'SDD12 MLOG', 'SDD21 GDEL', 'SDD21 MLOG', 'SDD21 POL1', 'SDD21 POL2',
                     'SDD11 MLOG', 'SDD22 MLOG', 'AV', 'CMRR1', 'CMRR2', 'Group Delay', 'S12_V']
        col = 1
        for item in firstline:
            sheet_ranges.cell(column=col, row=1, value=firstline[col-1])
            col = col + 1


    # Impedance parameters


    # Frequency sweep parameters


    # Measurements to be taken. See online VNA guide for more options
    measlist = ['SCC21 MLOG', 'SDC21 MLOG', 'SDD11 POL', 'SDD12 MLOG', 'SDD21 GDEL',
                'SDD21 MLOG', 'SDD21 POL', 'SDD11 MLOG', 'SDD22 MLOG']

    # Summary setup


    # header()
    instDict['Supply'].__SetEnable__(1)
    pluto = PlutoV0()
    pluto.connect(DUT1_Default=0x00, DUT2_Default=0x00)
    pluto.Set_Amp_Gain(SPI_sel=channel, GainValue='12dB')
    pluto.Set_Amp_Coupling(SPI_sel=channel, Coupling='OFF')
    pluto.Set_Amp_Pwr_Mod(SPI_sel=channel, PowerMode="Hi")
    pluto.Set_Amp_Enable(SPI_sel=channel, AmpEnable=True)
    pluto.Set_Amp_Trim(SPI_sel=channel, AmpTrimCode=15)
    # print pluto.GetReadDataRegisterValue(SPI_sel=channel)

    vocms = []


    VNAinit()

    # Main loop structure
    for temp in templist:
        if templist != [25]:
            setTemp(temp)
        fh.write('Temp = %d' % temp)
        fh.write('\n')
        for supply in supplylist:
            fh.write('Supply = %g\n' % supply)
            instDict['Supply'].__SetV__(supply)
            for vcom in vcomlist:
                if vcom != 'N/A':
                    instDict['Vcom'].__SetV__(vcom)
                    fh.write('Vcom = %g\n' % vcom)
                    time.sleep(0.2)
                for Pout in Poutlist:
                    fh.write('Pout != %g' % Pout)
                    print 'Pout not yet implemented'
                    for pwrmode in pwrmodelist:
                        fh.write('Power Mode = %s' % pwrmode)
                        pluto.Set_Amp_Pwr_Mod(SPI_sel=channel, PowerMode=pwrmode)
                        for gain in gainlist:
                            fh.write('Gain = %s\n' % gain)
                            pluto.Set_Amp_Gain(SPI_sel=channel, GainValue=gain)
                            # print pluto.GetReadDataRegisterValue(SPI_sel=channel)
                            getData()

    # Final actions: return to temperature, get execution time and close file
    if templist != [25]:
        setTemp(25)
        instDict['thermo'].__FlowOFF__()
    pluto.Set_Amp_Enable(SPI_sel=channel, AmpEnable=False)
    instDict['Supply'].__SetEnable__(0)
    instDict['NA'].__Output__(0)
    endTime = time.time() - startTime

    print 'Program executed in %d seconds.' % endTime

    fh.write('Execution Time = ,%d' % endTime)

    fh.close()
    wb.save(filename=summaryPath)

if __name__ == '__main__':
    path = 'C:\\Users\\bsulliv2\\Desktop\\Results\\PlutoV0\\SParam\\'
    summaryPath = 'C:\\Users\\bsulliv2\\Desktop\\Results\\PlutoV0\\PlutoSParamSummaryTemp.xlsx'

    Zin_diff = 100
    Zout_diff = 100
    avg = 16

    numPoints = 1000.0
    startFreq = 10e6
    endFreq = 10.01e9

    templist = [-40, 85, 25]
    vcomlist = ['N/A']
    supplylist = [3.3]
    gainlist = ['12dB', '20dB']
    pwrmodelist = ['Lo', 'Hi']
    # gainlist = ['12dB']
    # pwrmodelist = ['Lo']
    Poutlist = [-30]
    dut = 'V0B1'
    channel = 'A'

    instDict = InstInit()



    SParam()


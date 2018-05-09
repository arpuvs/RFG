# Author: Ben Sullivan
# Date: 11/06/2017

# Necessary module and instrument imports
import sys, visa, time
sys.path.append('../../Common')
sys.path.append('../../Common/FMB_USB_Python_Files')
from ADI_GPIB.AgilentN5181A import *
from ADI_GPIB.AgilentN9030A import *
from ADI_GPIB.AgilentN6705B import *
from ADI_GPIB.WatlowF4 import *
# from FMB import *
from openpyxl import *
from BSTest import *
from PlutoV0 import PlutoV0


# Main body of code, called by GUI
def HD23Main(path, supplyVlist, freqlist, vcomlist, templist, dutNumber, vcomEnable):

    def HD23():  # Main HD23 measurement function
        # Initializes analyzer
        instDict['SA'].__SetSpan__(span)
        instDict['SA'].__EnableAverage__(1)
        instDict['SA'].__SetAverage__(avg)
        instDict['SA'].__SetAutoBW__(0)
        instDict['SA'].__SetBW__(BW)

        # Initializes all measurement arrays
        fundamental = []
        second = []
        third = []
        Carrier = []
        SourceAmp = []
        secondNormal = []
        thirdNormal = []
        supplyI = []
        freqlistReal = []
        adjustedfreqlist = []

        for freqindex in freqlist:
            # Converts fmb dict frequency to float
            freq = fmbDict[freqindex].split()
            if freq[1] == 'MHz':
                freq = float(freq[0]) * 1e6
            elif freq[1] == 'GHz':
                freq = float(freq[0]) * 1e9
            freqlistReal.append(freq)

            # Sets up all equipment for given frequency
            # Filter.select_filter(freqindex)
            instDict['Source'].__SetFrequency__(freq/1e6)
            instDict['SA'].__Setfc__(freq)
            instDict['Source'].__SetOut_ON__()
            instDict['SA'].__SetMarkerFreq__(1, freq)
            time.sleep(0.1)
            instDict['SA'].__ClearAverage__()

            # Sets initial amplitude using measured output of dut
            instDict['SA'].__CheckStatus__(300)                     # Waits until averaging is complete
            time.sleep(0.5)
            instDict['SA'].__ClearAverage__()
            instDict['SA'].__CheckStatus__(300)

            if pksearchEnable:
                instDict['SA'].__PeakSearch__(1)
                time.sleep(0.1)
                freq = float(instDict['SA'].__GetMarkerFreq__(1))
                adjustedfreqlist.append(freq)
            else:
                adjustedfreqlist = freqlistReal

            carrierMag = float(instDict['SA'].__GetMarkerAmp__(1))  # Gets initial amplitude
            while abs(carrierMag - Pout) >= 0.1:                # Adjusts amplitude until it is within 0.1 dBm of desired
                sourceAmp = float(instDict['Source'].__GetLevel__())
                setAmp = sourceAmp + (Pout - carrierMag)
                if setAmp > maxPow:
                    raise Exception('Max source amplitude exceeded')
                instDict['Source'].__SetLevel__(setAmp)
                instDict['SA'].__ClearAverage__()
                instDict['SA'].__CheckStatus__(300)
                carrierMag = float(instDict['SA'].__GetMarkerAmp__(1))

            # Gets final amplitude
            Carrier.append(carrierMag)
            fundamental.append(carrierMag)
            SourceAmp.append(float(instDict['Source'].__GetLevel__()))

            # Makes second and third harmonic measurements
            instDict['SA'].__Setfc__(freq*2.0)
            instDict['SA'].__SetMarkerFreq__(1, freq*2.0)
            instDict['SA'].__ClearAverage__()
            instDict['SA'].__CheckStatus__(300)
            if pksearchEnable:
                instDict['SA'].__PeakSearch__(1)
                time.sleep(0.1)
            second.append(float(instDict['SA'].__GetMarkerAmp__(1)))
            instDict['SA'].__Setfc__(freq*3.0)
            instDict['SA'].__SetMarkerFreq__(1, freq*3.0)
            instDict['SA'].__ClearAverage__()
            instDict['SA'].__CheckStatus__(300)
            if pksearchEnable:
                instDict['SA'].__PeakSearch__(1)
                time.sleep(0.1)
            third.append(float(instDict['SA'].__GetMarkerAmp__(1)))
            supplyI.append(float(instDict['Supply'].__GetI__()))

        # Converts dBm measurements to dBc
        for j in range(len(second)):
            secondNormal.append(-(float(Carrier[j]) - float(second[j])))
            thirdNormal.append(-(float(Carrier[j]) - float(third[j])))

        # Appends measured data at end of specified xlsx file
        Acolumn = sheet_ranges['A']
        index = 0
        for row in range(len(Acolumn)+1, len(Acolumn) + len(adjustedfreqlist) + 1):
            data = [dutNumber, channel, temp, supplyV, vcom, Pout, pwrmode, gain, adjustedfreqlist[index], supplyI[index], Carrier[index],
                    second[index], third[index], secondNormal[index], thirdNormal[index]]
            for col in range(len(data)):
                sheet_ranges.cell(column=col+1, row=row, value=data[col])
            index = index + 1

    # Sets oven to setpoint and soaks dut for allotted time
    def setTemp(setpoint):
        Oven.__SetTemp__(setpoint)
        current = float(Oven.__GetTemp__())
        while (abs(current - setpoint) > 2):
            time.sleep(1)
            current = float(Oven.__GetTemp__())
        print '@ Temp %d' % setpoint
        soak = 300
        time.sleep(soak)
        return True

    # Sets analyzer attenuation settings (22dB was empirically chosen)
    startTime = time.time()
    instDict['SA'].__SetAutoAtten__(0)
    instDict['SA'].__SetAtten__(22)

    # Filter box frequency settings dictionary
    fmbDict = {1: "2.5 MHz", 2: "5 MHz", 3: "33 MHz", 4: "78 MHz",
              5: "120 MHz", 6: "225.3 MHz", 7: "350.3 MHz", 8: "500 MHz",
              9: "800.3 MHz", 10: "1 GHz", 11: "1.5 GHz", 12: "2 GHz",
              13: "2.5 GHz", 14: "3 GHz", 15: "3.5 GHz", 16: "4 GHz",
              17: "4.5 GHz", 18: "5 GHz", 19: "5.5 GHz", 20: "5.9 GHz",
              21: "Aux"}

    # Filter = FMB('COM3', fmbDict)  # Initializes filter box

    # Opens specified xlsx file. If no file exists a new spreadsheet is created.
    try:
        wb = load_workbook(filename=path)
        sheet_ranges = wb['Sheet1']
    except:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = 'Sheet1'
        sheet_ranges = wb['Sheet1']
        firstline = ['DUT', 'Channel', 'Temp', 'Supply', 'Vcom', 'Pout', 'PowerMode', 'Gain', 'Frequency', 'Current', 'Carrier(dB)',
                     'Second(dB)', 'Third(dB)', 'HD2(dBc)', 'HD3(dBc)']
        col = 1
        for item in firstline:
            sheet_ranges.cell(column=col, row=1, value=firstline[col-1])
            col = col + 1


    # Sets up main supply
    instDict['Supply'].__SetV__(supplies[0])
    instDict['Supply'].__SetI__(0.25)
    instDict['Supply'].__SetEnable__(1)
    pluto = PlutoV0()
    pluto.connect(DUT1_Default=0x00, DUT2_Default=0x00)
    pluto.Set_Amp_Gain(SPI_sel=channel, GainValue='12dB')
    pluto.Set_Amp_Coupling(SPI_sel=channel, Coupling='ON')
    pluto.Set_Amp_Pwr_Mod(SPI_sel=channel, PowerMode="Hi")
    pluto.Set_Amp_Enable(SPI_sel=channel, AmpEnable=True)
    pluto.Set_Amp_Trim(SPI_sel=channel, AmpTrimCode=15)
    if vcomEnable:
        instDict['Vcom'].__SetEnable__(1)

    # Main loop structure
    for temp in templist:
        if templist != [25]:
            setTemp(temp)                       # Sets DUT to temperature
        for supplyV in supplies:
            instDict['Supply'].__SetV__(supplyV)
            for vcom in vcomlist:
                instDict['Vcom'].__SetV__(vcom)  # Sets DUT to common mode voltage
                print 'Vcom = %g' % vcom
                for Pout in Poutlist:
                    for pwrmode in pwrmodelist:
                        pluto.Set_Amp_Pwr_Mod(SPI_sel=channel, PowerMode=pwrmode)
                        for gain in gainlist:
                            instDict['Source'].__SetLevel__(0)
                            pluto.Set_Amp_Gain(SPI_sel=channel, GainValue=gain)
                            print 'Gain = %s' % gain

                            instDict['Source'].__SetOut_OFF__()
                            print 'Aligning...'
                            # instDict['SA'].__CheckStatus__(300)
                            # instDict['SA'].__Align__()            # Aligns device with no input
                            # instDict['SA'].__CheckStatus__(300)
                            print 'Done'

                            HD23()                          # Runs main HD23 measurement

    # Returns oven to ambient temp, finds execution time and saves file\
    instDict['Source'].__SetOut_OFF__()
    pluto.Set_Amp_Enable(SPI_sel=channel, AmpEnable=False)
    instDict['Supply'].__SetEnable__(0)
    if vcomEnable:
        instDict['Vcom'].__SetEnable(0)
    if templist != [25]:
        Oven.__SetTemp__(25)
    print time.time()-startTime
    wb.save(filename=path)
    print 'Done!'


# Called if program run by itself
if __name__ == '__main__':
    # Sets all necessary variables
    path = 'C:\\Users\\bsulliv2\\Desktop\\Results\\PlutoV0\\V0HD23.xlsx'
    # freqlist = [100e6, 250e6, 500e6, 1.0e9, 1.5e9, 2.0e9, 2.5e9, 3.0e9, 3.5e9, 4.0e9]
    # freqs = [5, 6, 8, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20]
    freqs = [1, 5, 8, 10, 11, 12]
    # for i in range(20, 31):
    #     vcoms.append(i/10.0)
    # vcoms = [2.0, 2.5, 3.0]
    supplies = [3.3]
    vcoms = [2.5]
    temps = [25]
    vcomEnable = 0
    pksearchEnable = 1
    gainlist = ['12dB', '20dB']
    pwrmodelist = ['Lo', 'Hi']
    Poutlist = [-8]
    maxPow = 15
    span = 1e3
    avg = 10
    BW = 10

    dut = 'V0B3'
    channel = 'A'

    instDict = InstInit()




    # pluto.Set_Amp_Enable(SPI_sel=channel, AmpEnable=True)
    #
    # raw_input('Waiting...')

    HD23Main(path, supplies, freqs, vcoms, temps, dut, vcomEnable)  # Calls main program




# Author: Ben Sullivan
# Date: 10/27/2017

# Module and instrument imports
import sys, visa, time
sys.path.append('../../Common')
from ADI_GPIB.AgilentN5181A import *
from ADI_GPIB.AgilentN9030A import *
from ADI_GPIB.AgilentN6705B import *
from ADI_GPIB.WatlowF4 import *
from pushbullet import pushbullet
from PlutoV0 import PlutoV0
from BSTest import *
from openpyxl import *
from collections import OrderedDict
import csv

# Instrument initialization
startTime = time.time()

def IMD3main(path, freqlist, vcomlist, templist, dut):
    # Main IMD measurement program
    def IMD3():
        # Initializes spectrum analyzer for measurements
        instDict['SA'].__SetSpan__(10e3)
        instDict['SA'].__SetAverage__(50)

        # Initializes arrays
        lowerPeak = []
        higherPeak = []
        lowCarrier = []
        highCarrier = []
        higherSourceAmp = []
        lowerSourceAmp = []
        supplyI = []


        for freq in freqlist:
            # Set sources
            instDict['source1'].__SetState__(0)
            instDict['source2'].__SetState__(0)
            instDict['source1'].__SetFreq__(freq - delta/2)
            instDict['source2'].__SetFreq__(freq + delta/2)

            # Configure and measure low f Carrier
            instDict['SA'].__Setfc__(freq - delta/2)
            instDict['source1'].__SetState__(1)
            instDict['SA'].__ClearAverage__()
            instDict['SA'].__CheckStatus__(300)   # Waits for averaging
            if pkSearch:
                instDict['SA'].__PeakSearch__(1)
            else:
                instDict['SA'].__SetMarkerFreq__(1, freq - delta/2)
            carrierMag = float(instDict['SA'].__GetMarkerAmp__(1))  # Measures initial amplitude
            while abs(carrierMag - Pout) >= 0.1:  # Adjusts until carrier mag is within 0.1dBm of desired amplitude
                sourceAmp = float(instDict['source1'].__GetAmp__())     # Gets amplitude
                setAmp = sourceAmp + (Pout - carrierMag)      # Adjusts amplitude
                if setAmp >= 15:    # Raises flag if max amplitude exceeded
                    raise Exception('Amplitude too high, check configuration')
                instDict['source1'].__SetAmp__(setAmp)      # Sets new amplitude
                instDict['SA'].__ClearAverage__()
                instDict['SA'].__CheckStatus__(300)
                carrierMag = float(instDict['SA'].__GetMarkerAmp__(1))    # Gets new amplitude

            # Records measured source and carrier values
            lowCarrier.append(float(instDict['SA'].__GetMarkerAmp__(1)))
            lowerSourceAmp.append(float(instDict['source1'].__GetAmp__()))

            # Repeat for high f carrier - same process as above
            instDict['source1'].__SetState__(0)
            instDict['SA'].__Setfc__(freq + delta/2)
            instDict['source2'].__SetState__(1)
            instDict['SA'].__ClearAverage__()
            instDict['SA'].__CheckStatus__(300)
            if pkSearch == True:
                instDict['SA'].__PeakSearch__(1)
            else:
                instDict['SA'].__SetMarkerFreq__(1, freq + delta/2)
            carrierMag = float(instDict['SA'].__GetMarkerAmp__(1))
            while abs(carrierMag - Pout) >= 0.1:
                sourceAmp = float(instDict['source2'].__GetAmp__())
                setAmp = sourceAmp + (Pout - carrierMag)
                if setAmp >= 15:
                    raise Exception('Amplitude too high, check configuration')
                instDict['source2'].__SetAmp__(setAmp)
                instDict['SA'].__ClearAverage__()
                instDict['SA'].__CheckStatus__(300)
                carrierMag = float(instDict['SA'].__GetMarkerAmp__(1))
            highCarrier.append(float(instDict['SA'].__GetMarkerAmp__(1)))
            higherSourceAmp.append(float(instDict['source2'].__GetAmp__()))

            # Measure lower IMD3
            instDict['source1'].__SetState__(1)
            instDict['SA'].__Setfc__(freq - (1.5*delta))
            instDict['SA'].__ClearAverage__()
            instDict['SA'].__CheckStatus__(300)
            if pkSearch == True:
                instDict['SA'].__PeakSearch__(1)
            else:
                instDict['SA'].__SetMarkerFreq__(1, freq - (1.5*delta))
            # raw_input('Waiting...')
            lowerPeak.append(float(instDict['SA'].__GetMarkerAmp__(1)))

            # Measure higher IMD3
            instDict['SA'].__Setfc__(freq + (1.5*delta))
            instDict['SA'].__ClearAverage__()
            instDict['SA'].__CheckStatus__(300)
            if pkSearch == True:
                instDict['SA'].__PeakSearch__(1)
            else:
                instDict['SA'].__SetMarkerFreq__(1, freq + (1.5*delta))
            higherPeak.append(float(instDict['SA'].__GetMarkerAmp__(1)))
            supplyI.append(float(instDict['Supply'].__GetI__()))

        leftNormal = []
        rightNormal = []
        # Converts dBm measurements to dBc
        for j in range(len(lowerPeak)):
            leftNormal.append(-(float(lowCarrier[j]) - float(lowerPeak[j])))
            rightNormal.append(-(float(highCarrier[j]) - float(higherPeak[j])))

        for row in range(len(Acolumn)+1, len(Acolumn) + len(adjustedfreqlist) + 1):
            data = [dut, channel, temp, supplyV, vcom, atten, freqlist[index], supplyI[index], lowerPeak[index],
                    higherPeak[index], highCarrier[index], lowCarrier[index], leftNormal[index], rightNormal[index]]
            for col in range(len(data)):
                sheet_ranges.cell(column=col+1, row=row, value=data[col])
            index = index + 1

        # Writes all results to output file
        # fh.write('Test %d' % i)
        # fh.write('\n')
        fh.write('Frequency:,')
        fh.write(str(freqlist).strip('[]'))
        fh.write('\n')
        fh.write('Supply Current:,')
        fh.write(str(supplyI).strip('[]'))
        fh.write('\n')
        fh.write('Low left:,')
        fh.write(str(lowerPeak).strip('[]'))
        fh.write('\n')
        fh.write('Low right:,')
        fh.write(str(higherPeak).strip('[]'))
        fh.write('\n')
        fh.write('High left:,')
        fh.write(str(highCarrier).strip('[]'))
        fh.write('\n')
        fh.write('High right:,')
        fh.write(str(lowCarrier).strip('[]'))
        fh.write('\n')
        fh.write('Left dBc:,')
        fh.write(str(leftNormal).strip('[]'))
        fh.write('\n')
        fh.write('Right dBc:,')
        fh.write(str(rightNormal).strip('[]'))
        fh.write('\n')

    def IMD2():
        # Initializes spectrum analyzer for measurements
        instDict['SA'].__SetSpan__(10e3)
        instDict['SA'].__SetAverage__(50)

        # Initializes arrays
        lowerPeak = []
        higherPeak = []
        lowCarrier = []
        highCarrier = []
        higherSourceAmp = []
        lowerSourceAmp = []
        supplyI = []

        for freq in freqlist:
            # Set sources
            instDict['source1'].__SetState__(0)
            instDict['source2'].__SetState__(0)
            instDict['source1'].__SetFreq__(freq - delta2 / 2)
            instDict['source2'].__SetFreq__(freq + delta2 / 2)

            # Configure and measure low f Carrier
            instDict['SA'].__Setfc__(freq - delta2 / 2)
            instDict['source1'].__SetState__(1)
            instDict['SA'].__ClearAverage__()
            instDict['SA'].__CheckStatus__(300)  # Waits for averaging
            if pkSearch:
                instDict['SA'].__PeakSearch__(1)
            else:
                instDict['SA'].__SetMarkerFreq__(1, freq - delta2 / 2)
            carrierMag = float(instDict['SA'].__GetMarkerAmp__(1))  # Measures initial amplitude
            while abs(carrierMag - Pout) >= 0.1:  # Adjusts until carrier mag is within 0.1dBm of desired amplitude
                sourceAmp = float(instDict['source1'].__GetAmp__())  # Gets amplitude
                setAmp = sourceAmp + (Pout - carrierMag)  # Adjusts amplitude
                if setAmp >= 15:  # Raises flag if max amplitude exceeded
                    raise Exception('Amplitude too high, check configuration')
                instDict['source1'].__SetAmp__(setAmp)  # Sets new amplitude
                instDict['SA'].__ClearAverage__()
                instDict['SA'].__CheckStatus__(300)
                carrierMag = float(instDict['SA'].__GetMarkerAmp__(1))  # Gets new amplitude

            # Records measured source and carrier values
            lowCarrier.append(float(instDict['SA'].__GetMarkerAmp__(1)))
            lowerSourceAmp.append(float(instDict['source1'].__GetAmp__()))

            # Repeat for high f carrier - same process as above
            instDict['source1'].__SetState__(0)
            instDict['SA'].__Setfc__(freq + delta2 / 2)
            instDict['source2'].__SetState__(1)
            instDict['SA'].__ClearAverage__()
            instDict['SA'].__CheckStatus__(300)
            if pkSearch == True:
                instDict['SA'].__PeakSearch__(1)
            else:
                instDict['SA'].__SetMarkerFreq__(1, freq + delta2 / 2)
            carrierMag = float(instDict['SA'].__GetMarkerAmp__(1))
            while abs(carrierMag - Pout) >= 0.1:
                sourceAmp = float(instDict['source2'].__GetAmp__())
                setAmp = sourceAmp + (Pout - carrierMag)
                if setAmp >= 15:
                    raise Exception('Amplitude too high, check configuration')
                instDict['source2'].__SetAmp__(setAmp)
                instDict['SA'].__ClearAverage__()
                instDict['SA'].__CheckStatus__(300)
                carrierMag = float(instDict['SA'].__GetMarkerAmp__(1))
            highCarrier.append(float(instDict['SA'].__GetMarkerAmp__(1)))
            higherSourceAmp.append(float(instDict['source2'].__GetAmp__()))

            # Measure lower IMD2
            instDict['source1'].__SetState__(1)
            instDict['SA'].__Setfc__(1, delta2)
            instDict['SA'].__ClearAverage__()
            instDict['SA'].__CheckStatus__(300)
            if pkSearch == True:
                instDict['SA'].__PeakSearch__(1)
            else:
                instDict['SA'].__SetMarkerFreq__(1, 2*freq)
            # raw_input('Waiting...')
            lowerPeak.append(float(instDict['SA'].__GetMarkerAmp__(1)))

            # Measure higher IMD2
            instDict['SA'].__Setfc__(freq + (1.5 * delta2))
            instDict['SA'].__ClearAverage__()
            instDict['SA'].__CheckStatus__(300)
            if pkSearch == True:
                instDict['SA'].__PeakSearch__(1)
            else:
                instDict['SA'].__SetMarkerFreq__(1, freq + (1.5 * delta2))
            higherPeak.append(float(instDict['SA'].__GetMarkerAmp__(1)))
            supplyI.append(float(instDict['Supply'].__GetI__()))

        leftNormal = []
        rightNormal = []
        # Converts dBm measurements to dBc
        for j in range(len(lowerPeak)):
            leftNormal.append(-(float(lowCarrier[j]) - float(lowerPeak[j])))
            rightNormal.append(-(float(highCarrier[j]) - float(higherPeak[j])))

        for row in range(len(Acolumn)+1, len(Acolumn) + len(adjustedfreqlist) + 1):
            data = [lowerPeak[index], higherPeak[index], highCarrier[index], lowCarrier[index],
                    leftNormal[index], rightNormal[index]]
            for col in range(len(data)):
                sheet_ranges.cell(column=col+1, row=row, value=data[col])
            index = index + 1

        # Writes all results to output file
        # fh.write('Test %d' % i)
        # fh.write('\n')
        fh.write('Frequency:,')
        fh.write(str(freqlist).strip('[]'))
        fh.write('\n')
        fh.write('Supply Current:,')
        fh.write(str(supplyI).strip('[]'))
        fh.write('\n')
        fh.write('Low left:,')
        fh.write(str(lowerPeak).strip('[]'))
        fh.write('\n')
        fh.write('Low right:,')
        fh.write(str(higherPeak).strip('[]'))
        fh.write('\n')
        fh.write('High left:,')
        fh.write(str(highCarrier).strip('[]'))
        fh.write('\n')
        fh.write('High right:,')
        fh.write(str(lowCarrier).strip('[]'))
        fh.write('\n')
        fh.write('Left dBc:,')
        fh.write(str(leftNormal).strip('[]'))
        fh.write('\n')
        fh.write('Right dBc:,')
        fh.write(str(rightNormal).strip('[]'))
        fh.write('\n')

        # Print first line of file
    def header():
        test = 'IMD3'
        equipment = 'N5181A N9030A BAL0026 BAL006'
        # supplyV = instDict['Supply'].__MeasP25V__()
        supplyV = 'Temp'
        print supplyV
        # supplyI = instDict['Supply'].__MeasP25I__()
        supplyI = 'Temp'
        print supplyI
        balun = 'INB: 0-VIN OUTB 0-VOP'
        header = (dut, date, test, equipment, supplyV, supplyI, balun)
        header = str(header).strip('()')
        fh.write(header)
        fh.write('\n')

    # Sets oven temperature to setpoint and soaks DUT
    def setTemp(temperature):
            # Could add function to turn off supplies when ramping up per Greg's suggestion
            instDict['thermo'].__SetupDUTMode__(HighTemp=145, LowTemp=-70, SensorType='K', TestTime=230)
            instDict['thermo'].__SetDutDtype__(1)
            instDict['thermo'].__SetDutThermalConst__(100)
            instDict['thermo'].__EnableDUTMode__(SensorType='K')
            off = False
            instDict['thermo'].__SetTemp__(temperature)
            instDict['thermo'].__FlowON__()
            instDict['thermo'].__MoveArmDown__()
            instDict['thermo'].__EnableDUTMode__('K')
            time.sleep(5)
            measTemp = instDict['thermo'].__GetDutTemperature__()
            # if temperature > measTemp:
            #     powerDown(instDict)
            #     off = True
            while (abs(temperature - measTemp) > 5):
                time.sleep(10)
                measTemp = instDict['thermo'].__GetDutTemperature__()
                print measTemp
            print 'At temp'
            print 'Soaking...'
            time.sleep(300)
            # if off:
            #     powerUp(instDict)
            #     # chip = Linus(bridge_device='aardvark', linus_rev=2)


    # Initializes and opens output file
    date = time.ctime(time.time())
    date = date.replace(':', '-')
    fh = open(path + 'Intermod_Dist_' + date + '.csv', 'w')

    header()    # Prints first line of file
    # results = OrderedDict()
    try:
        wb = load_workbook(filename=path)
        sheet_ranges = wb['Sheet1']
    except:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = 'Sheet1'
        sheet_ranges = wb['Sheet1']
        firstline = ['DUT', 'Channel', 'Temp', 'Supply', 'Vcom', 'Power Mode', 'Gain', 'Pout', 'Frequency', 'Current',
                     'Low Left3(dB)', 'Low Right3(dB)', 'High Left3(dB)', 'High Right3(dB)', 'IMD3LO(dBc)',
                     'IMD3HI(dbC)', 'Low Left2(dB)', 'Low Right2(dB)', 'High Left2(dB)', 'High Right2(dB)',
                     'IMD2LO(dBc)', 'IMD2HI(dbC)']
        col = 1
        for item in firstline:
            sheet_ranges.cell(column=col, row=1, value=firstline[col-1])
            col = col + 1

    # Configure supplies
    # Vsupply = Ch1, Ven = Ch2, Vcom = Ch3
    instDict['Supply'].__SetV__(5)
    instDict['Supply'].__SetI__(0.25)


    # balunList = ('standard', 'input flipped', 'both flipped', 'output flipped')
    # balunList = ('S')
    pluto = PlutoV0()
    pluto.connect(DUT1_Default=0x00, DUT2_Default=0x00)
    pluto.Set_Amp_Gain(SPI_sel=channel, GainValue='20dB')
    pluto.Set_Amp_Coupling(SPI_sel=channel, Coupling='ON')
    pluto.Set_Amp_Pwr_Mod(SPI_sel=channel, PowerMode="Hi")
    pluto.Set_Amp_Enable(SPI_sel=channel, AmpEnable=True)
    pluto.Set_Amp_Trim(SPI_sel=channel, AmpTrimCode=15)
    for temp in templist:
    # for balun in balunList:
        if templist != [25]:
            setTemp(temp)
        # fh.write('Balun config = %s' % balun)
        fh.write('Temp = %d' % temp)
        fh.write('\n')
        for vcom in vcomlist:
            instDict['vcom'].__SetV__(vcom)  # Sets DUT to common mode voltage

            instDict['source1'].__SetState__(0)
            instDict['source2'].__SetState__(0)
            if align:
                print 'Aligning...'
                instDict['SA'].__CheckStatus__(300)
                instDict['SA'].__Align__()  # Aligns device with no input
                instDict['SA'].__CheckStatus__(300)
            print 'Done'

            print 'Vcom = %g' % vcom
            fh.write('Vcom = %g\n' % vcom)
            # raw_input('Configure Bal uns to : %s' % balun)
            #     for i in range(1):
            for Pout in Poutlist:
                fh.write('Pout = %g' % Pout)
                for pwrmode in pwrmodelist:
                    fh.write('PowerMode = %s' % pwrmode)
                    pluto.Set_Amp_Pwr_Mod(SPI_sel=channel, PowerMode=pwrmode)
                    for gain in gainlist:
                        fh.write('Gain = %s\n' % gain)
                        pluto.Set_Amp_Gain(SPI_sel=channel, GainValue=gain)

                        IMD3()  # Runs main IMD measurements

    # Final clean up actions
    pluto.Set_Amp_Enable(SPI_sel=channel, AmpEnable=False)
    if templist != [25]:
        instDict['thermo'].__SetTemp__(25)
    elapsedtime = time.time() - startTime
    print elapsedtime
    fh.write('Execution time (s) = %g' % (elapsedtime))

    key = 'o.0fWIoYany1oZxKkpg4pGHWvJUYqVHXPW'
    p = pushbullet.PushBullet(key)
    devices = p.getDevices()
    p.pushNote(devices[0]['iden'], 'Demo Test Done', '%d' % elapsedtime)

if __name__ == '__main__':
    pkSearch = True
    path = 'C:\\Users\\bsulliv2\\Desktop\\Results\\PlutoV0\\SA_IMD\\'
    summarypath = 'C:\\Users\\bsulliv2\\Desktop\\Results\\Pluto\\V0\\IMD.xlsx'
    # freqs = [100e6, 250e6, 500e6, 1.0e9, 1.5e9, 2.0e9, 2.5e9, 3.0e9, 3.5e9, 4.0e9, 4.5e9, 5.0e9, 5.5e9, 5.9e9]
    freqs = [100e6, 250e6, 500e6, 1.0e9, 1.5e9, 2.0e9]
    # freqs = [6e9]
    # vcoms = []
    # for i in range(20, 31):
    #     vcoms.append(i/10.0)
    # vcoms = [2.0, 2.5, 3.0]
    vcoms = [1.2]
    # temps = [25, 85, -40]
    delta = 2e6
    temps = [25]
    dut = 'V0B2 A + 10dB pads Output referred'
    channel = 'A'
    align = True

    gainlist = ['12dB', '20dB']
    pwrmodelist = ['Lo', 'Hi']
    Poutlist = [-12]

    instDict = InstInit()

    IMD3main(path, freqs, vcoms, temps, dut)  # Calls main program


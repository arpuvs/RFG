# Author: Ben Sullivan
# Date: 12/19/2017

import sys, visa, time, math, cmath
sys.path.append('../../Common')

# Instrumnet imports and initialization
# from ADI_GPIB.WatlowF4 import *
# from ADI_GPIB.E3631A import *
# from ADI_GPIB.KeysightN5242A import *
from openpyxl import *
from BSTest import *
from PlutoV0 import PlutoV0
from pushbullet import pushbullet


# VNA = KeysightN5424A(17)
# Oven = WatlowF4(4)
# Supply = E3631A(23)
# vcomSupply = E3631A(15)

# Sets all necessary VNA parameters and adds all specified measurements
def IMD():
    def VNAinit():
        instDict['NA'].__Preset__('full')
        instDict['NA'].__AddWindow__(1)
        trace = 1

        # Adds measurements from supplied list
        for meas in measlist:
            instDict['NA'].__AddMeas__(1, meas, 'Swept IMD', meas)
            instDict['NA'].__AddTrace__(1, trace, meas)
            trace = trace + 1
        instDict['NA'].__SetStartf__(1, startFreq)
        instDict['NA'].__SetStopf__(1, endFreq)
        instDict['NA'].__SetNumPoints__(1, numPoints)
        instDict['NA'].__SetAvg__(1, avg)
        instDict['NA'].__SetAutoTime__(1, True)
        instDict['NA'].__EnableAvg__(1, True)
        # instDict['NA'].__IMDPower__(1, Pout)
        instDict['NA'].__RecallCal__('Pluto_IMD2')
        instDict['NA'].__SetAttenuation__(0)
        instDict['NA'].__SetSourceAttenuation__(1, 2, 0)
        instDict['NA'].__SetStopf__(1, endFreq)
        instDict['NA'].__SetNumPoints__(1, numPoints)
        instDict['NA'].__SetIMDIFBW__(1, ProductIFBW)
        instDict['NA'].__SetMainIFBW__(1, MainIFBW)
        instDict['NA'].__IMDPowType__(1, 'OUTPUT')
        instDict['NA'].__IMDPower__(1, Pout)
        instDict['NA'].__SetIMDDelta__(1, float(delta))
        instDict['NA'].__EnableAvg__(1, True)
        instDict['NA'].__SetAvg__(1, avg)

    # Sets oven to specified temperature and soaks
    def setTemp(temperature):
        # Could add function to turn off supplies when ramping up per Greg's suggestion
        instDict['thermo'].__SetupDUTMode__(HighTemp=145, LowTemp=-70, SensorType='K', TestTime=230)
        instDict['thermo'].__SetDutDtype__(1)
        instDict['thermo'].__SetDutThermalConst__(100)
        instDict['thermo'].__EnableDUTMode__(SensorType='K')
        off = False
        instDict['thermo'].__SetTemp__(temperature)
        instDict['thermo'].__FlowON__()
        instDict['thermo'].__MoveArmDown__()
        instDict['thermo'].__EnableDUTMode__('K')
        time.sleep(5)
        measTemp = instDict['thermo'].__GetDutTemperature__()
        # if temperature > measTemp:
        #     powerDown(instDict)
        #     off = True
        while (abs(temperature - measTemp) > 5):
            time.sleep(10)
            measTemp = instDict['thermo'].__GetDutTemperature__()
            print measTemp
        print 'At temp'
        print 'Soaking...'
        time.sleep(300)
        # if off:
        #     powerUp(instDict)
        #     # chip = Linus(bridge_device='aardvark', linus_rev=2)

    # Retrieves measured data from VNA and prints to file
    def getData():
        readDict = {}   # Dictionary for results to be stored in
        instDict['NA'].__EnableAvg__(1, False)
        instDict['NA'].__EnableAvg__(1, True)

        time.sleep(20)  # Necessary to let one sweep finish. Otherwise FinishAvg function does not work
        print 'Done Sleeping'
        instDict['NA'].__FinishAvg__(1, 600)   # Pauses execution until VNA is finished averaging

        # Retrieve data from VNA
        for meas in measlist:
            instDict['NA'].__SetActiveTrace__(1, meas)
            ans = instDict['NA'].__GetData__(1)
            ans = ans.split(',')
            # Converts received data from unicode to numeric
            for val in range(len(ans)):
                ans[val] = float(ans[val])
            readDict[meas] = ans

        # Retrieves swept frequency points
        freqlist = instDict['NA'].__GetFreq__(1)
        # Writes all data to file
        fh.write('Frequency,')
        fh.write(str(freqlist).strip('[]'))
        for meas in measlist:
            fh.write(meas)
            fh.write(',')
            fh.write(str(readDict[meas]).strip('[]'))
            fh.write('\n')
        fh.write('\n')
        return readDict

    # Prints first line of file
    def header():
        dut = '3-4 ChA'
        test = 'PNA-X IMD'
        equipment = 'BAL0026 6dB in and out '
        header = (dut, date, test, equipment)
        header = str(header).strip('()')
        fh.write(header)
        fh.write('\n')

    def summary(freqlist, readDict2, readDict3):
        # columns = [dut, temp, supply, vcom, atten, freqlist, scc21]

        Acolumn = sheet_ranges['A']
        index = 0
        for row in range(len(Acolumn) + 1, len(Acolumn) + 49):
            data = [dut, temp, supply, vcom, MainIFBW, ProductIFBW, Pout, pwrmode, gain, freqlist[index], readDict2['PwrMainHi'][index],
                    readDict2['PwrMainLo'][index], readDict2['IM2HI'][index], readDict2['IM2LO'][index],
                    readDict2['PwrMainIN'][index], readDict2['OIP2LO'][index], readDict2['OIP2HI'][index],
                    readDict3['PwrMainHi'][index], readDict3['PwrMainHi'][index], readDict3['IM3HI'][index],
                    readDict3['IM3LO'][index], readDict3['PwrMainIN'][index], readDict3['OIP3LO'][index],
                    readDict3['OIP3HI'][index]]
            for col in range(len(data)):
                sheet_ranges.cell(column=col + 1, row=row, value=data[col])
            index = index + int(numPoints / 50)

        index = len(freqlist) - 1
        data = [dut, temp, supply, vcom, MainIFBW, ProductIFBW, Pout, pwrmode, gain, freqlist[index], readDict2['PwrMainHi'][index],
                readDict2['PwrMainLo'][index], readDict2['IM2HI'][index], readDict2['IM2LO'][index],
                readDict2['PwrMainIN'][index], readDict2['OIP2LO'][index], readDict2['OIP2HI'][index],
                readDict3['PwrMainHi'][index], readDict3['PwrMainHi'][index], readDict3['IM3HI'][index],
                readDict3['IM3LO'][index], readDict3['PwrMainIN'][index], readDict3['OIP3LO'][index],
                readDict3['OIP3HI'][index]]
        for col in range(len(data)):
            sheet_ranges.cell(column=col + 1, row=row + 1, value=data[col])


    startTime = time.time()

    date = time.ctime(time.time())
    date = date.replace(':', '-')

    fh = open(path + 'VNA_IMD' + date + '.csv', 'w')    # Creates csv file

    try:
        wb = load_workbook(filename=summaryPath)
        sheet_ranges = wb['Sheet1']
    except:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = 'Sheet1'
        sheet_ranges = wb['Sheet1']
        firstline = ['DUT', 'Temp', 'Supply', 'Vcom', 'MainIFBW', 'ProductIFBW', 'Pout', 'PowerMode', 'Gain', 'Frequency', 'PwrMain2Hi',
                     'PwrMain2Lo', 'IM2HI', 'IM2LO', 'PwrMainIn2', 'OIP2LO', 'OIP2HI', 'PwrMain3Hi', 'PwrMain3Lo',
                     'IM3HI', 'IM3LO', 'PwrMainIn3', 'OIP3LO', 'OIP3HI']
        col = 1
        for item in firstline:
            sheet_ranges.cell(column=col, row=1, value=firstline[col-1])
            col = col + 1

    # Measurements to be taken. See online VNA guide for more options
    measlist3 = ['PwrMainHi', 'PwrMainLo', 'IM3HI', 'IM3LO', 'PwrMainIN', 'OIP3LO', 'OIP3HI']
    measlist2 = ['PwrMainHi', 'PwrMainLo', 'IM2HI', 'IM2LO', 'PwrMainIN', 'OIP2LO', 'OIP2HI']

    header()
    instDict['Supply'].__SetEnable__(1)
    pluto = PlutoV0()
    pluto.connect(DUT1_Default=0x00, DUT2_Default=0x00)
    pluto.Set_Amp_Gain(SPI_sel=channel, GainValue='20dB')
    pluto.Set_Amp_Coupling(SPI_sel=channel, Coupling='ON')
    pluto.Set_Amp_Pwr_Mod(SPI_sel=channel, PowerMode="Hi")
    pluto.Set_Amp_Enable(SPI_sel=channel, AmpEnable=True)
    pluto.Set_Amp_Trim(SPI_sel=channel, AmpTrimCode=15)
    # VNAinit()

    # Main loop structure
    for temp in templist:
        if templist != [25]:
            setTemp(temp)
        fh.write('Temp = %d' % temp)
        fh.write('\n')
        for supply in supplylist:
            instDict['Supply'].__SetV__(supply)
            fh.write('Supply = %g\n' % supply)
            for vcom in vcomlist:
                if vcom != 'N/A':
                    instDict['vcom'].__SetV__(vcom)
                    instDict['vcom'].__SetEnable__(1)
                    fh.write('Vcom = %g\n' % vcom)
                    time.sleep(0.2)
                for Pout in Poutlist:
                    fh.write('Pout = %g' % Pout)
                    for pwrmode in pwrmodelist:
                        fh.write('PowerMode = %s' % pwrmode)
                        pluto.Set_Amp_Pwr_Mod(SPI_sel=channel, PowerMode=pwrmode)
                        for gain in gainlist:
                            fh.write('Gain = %s\n' % gain)
                            pluto.Set_Amp_Gain(SPI_sel=channel, GainValue=gain)
                            measlist = measlist2
                            delta = im2delta
                            VNAinit()
                            readDict2 = getData()
                            measlist = measlist3
                            delta = im3delta
                            VNAinit()
                            readDict3 = getData()
                            freqlist = instDict['NA'].__GetFreq__(1)
                            freqlist = freqlist.strip('[]')
                            freqlist = freqlist.split(',')
                            newfreqlist = []
                            for item in freqlist:
                                item = float(item)
                                newfreqlist.append(item)
                            summary(newfreqlist, readDict2, readDict3)

    # Final actions: return to temperature, get execution time and close file
    if templist != [25]:
        setTemp(25)
        instDict['thermo'].__FlowOFF__()
    pluto.Set_Amp_Enable(SPI_sel=channel, AmpEnable=False)
    instDict['Supply'].__SetEnable__(0)
    instDict['NA'].__Output__(0)
    endTime = time.time() - startTime

    print 'Program executed in %d seconds.' % endTime

    fh.write('Execution Time = ,%d' % endTime)

    fh.close()
    wb.save(filename=summaryPath)

    key = 'o.0fWIoYany1oZxKkpg4pGHWvJUYqVHXPW'
    p = pushbullet.PushBullet(key)
    devices = p.getDevices()
    p.pushNote(devices[0]['iden'], 'IMD Test Done', '%d' % endTime)

if __name__ == '__main__':
    path = 'C:\\Users\\bsulliv2\\Desktop\\Results\\PlutoV0\\IMD\\'
    summaryPath = 'C:\\Users\\bsulliv2\\Desktop\\Results\\PlutoV0\\PlutoIMDSummaryPoutTest.xlsx'

    avg = 3

    numPoints = 201
    startFreq = 10.5e6
    endFreq = 2.0105e9
    # Pout = -6

    im2delta = '10e6'
    im3delta = '2e6'
    templist = [25]
    # templist = [-40, 85, 25]
    vcomlist = [0.9]
    supplylist = [3.3]
    gainlist = ['12dB', '20dB']
    pwrmodelist = ['Lo', 'Hi']
    Poutlist = [-12]
    ProductIFBW = 10
    MainIFBW = 1000
    dut = 'V0B2 A + 10dB pads Output referred'
    channel = 'A'

    instDict = InstInit()



    IMD()


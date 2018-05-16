import sys, visa, time, math, cmath
sys.path.append('../../Common')

# Instrumnet imports and initialization
# from ADI_GPIB.WatlowF4 import *
# from ADI_GPIB.E3631A import *
# from ADI_GPIB.KeysightN5242A import *
from openpyxl import *
from BSTest import *
from PlutoV0 import PlutoV0


def NF():
    def VNAinit():
        instDict['NA'].__Preset__('full')
        instDict['NA'].__AddWindow__(1)
        trace = 1

        for meas in measlist:
            instDict['NA'].__AddMeas__(1, meas, 'Noise Figure Cold Source', meas)
            instDict['NA'].__AddTrace__(1, trace, meas)
            trace = trace + 1
        instDict['NA'].__SetStartf__(1, startFreq)
        instDict['NA'].__SetStopf__(1, endFreq)
        instDict['NA'].__SetNumPoints__(1, numPoints)
        # instDict['NA'].__SetAutoTime__(1, True)
        # instDict['NA'].__SetAttenuation__(10)
        # instDict['NA'].__SetSourceAttenuation__(1, 2, 0)
        instDict['NA'].__RecallCal__('Pluto_NF')
        instDict['NA'].__PortPower__(1, 1, P1)
        instDict['NA'].__PortPower__(1, 1, P2)
        instDict['NA'].__NoiseBW__(noiseBW)
        # instDict['MA'].__Temperature__(temp + 273)
        instDict['NA'].__EnableAvg__(1, True)
        instDict['NA'].__SetAvg__(1, avg)

    def setTemp(temperature):
            # Could add function to turn off supplies when ramping up per Greg's suggestion
            instDict['thermo'].__SetupDUTMode__(HighTemp=145, LowTemp=-70, SensorType='K', TestTime=230)
            instDict['thermo'].__SetDutDtype__(1)
            instDict['thermo'].__SetDutThermalConst__(100)
            instDict['thermo'].__EnableDUTMode__(SensorType='K')
            off = False
            instDict['thermo'].__SetTemp__(temperature)
            instDict['thermo'].__FlowON__()
            instDict['thermo'].__MoveArmDown__()
            instDict['thermo'].__EnableDUTMode__('K')
            time.sleep(5)
            measTemp = instDict['thermo'].__GetDutTemperature__()
            # if temperature > measTemp:
            #     powerDown(instDict)
            #     off = True
            while (abs(temperature - measTemp) > 5):
                time.sleep(10)
                measTemp = instDict['thermo'].__GetDutTemperature__()
                print measTemp
            print 'At temp'
            print 'Soaking...'
            time.sleep(300)
            # if off:
            #     powerUp(instDict)
            #     # chip = Linus(bridge_device='aardvark', linus_rev=2)

    def header():
        test = 'PNA-X NF'
        equipment = 'BAL0026 6dB in and out '
        header = (dut, date, test, equipment)
        header = str(header).strip('()')
        fh.write(header)
        fh.write('\n')

    def getData():
        readDict = {}   # Dictionary for results to be stored in
        time.sleep(5)  # Necessary to let one sweep finish. Otherwise FinishAvg function does not work
        instDict['NA'].__FinishAvg__(1, 600)   # Pauses execution until VNA is finished averaging

        # Retrieve data from VNA
        for meas in measlist:
            instDict['NA'].__SetActiveTrace__(1, meas)
            ans = instDict['NA'].__GetData__(1)
            ans = ans.split(',')
            # Converts received data from unicode to numeric
            for val in range(len(ans)):
                ans[val] = float(ans[val])
            # Polar measurements return real and imaginary data that must be separated
            readDict[meas] = ans

        # Retrieves and converts swept frequency points
        freqlist = instDict['NA'].__GetFreq__(1)
        freqlist = freqlist.split(',')
        for val in range(len(freqlist)):
            freqlist[val] = float(freqlist[val])

        # Writes all data to file
        fh.write('Frequency,')
        fh.write(str(freqlist).strip('[]'))
        fh.write('\n')
        for meas in measlist:
            fh.write(meas)
            fh.write(',')
            fh.write(str(readDict[meas]).strip('[]'))
            fh.write('\n')
        summary(freqlist, readDict)

    def summary(freqlist, readDict):

        Acolumn = sheet_ranges['A']
        index = 0
        for row in range(len(Acolumn) + 1, len(Acolumn) + 50):
            data = [dut, temp, supply, vcom, P2, pwrmode, gain, freqlist[index], readDict['NF'][index],
                    readDict['R1_1'][index], readDict['R2_2'][index], readDict['S21'][index]]
            for col in range(len(data)):
                sheet_ranges.cell(column=col + 1, row=row, value=data[col])
            index = index + int(numPoints / 50)

        index = len(freqlist) - 1
        data = [dut, temp, supply, vcom, P2, pwrmode, gain, freqlist[index], readDict['NF'][index],
                readDict['R1_1'][index], readDict['R2_2'][index], readDict['S21'][index]]
        for col in range(len(data)):
            sheet_ranges.cell(column=col + 1, row=row + 1, value=data[col])

    startTime = time.time()

    date = time.ctime(time.time())
    date = date.replace(':', '-')

    fh = open(path + 'VNA_IMD' + date + '.csv', 'w')    # Creates csv file

    try:
        wb = load_workbook(filename=summaryPath)
        sheet_ranges = wb['Sheet1']
    except:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = 'Sheet1'
        sheet_ranges = wb['Sheet1']
        firstline = ['DUT', 'Temp', 'Supply', 'Vcom', 'Pout', 'PowerMode', 'Gain', 'Frequency', 'Noise Figure',
                     'Port1 Power', 'Port2Power', 'S21']
        col = 1
        for item in firstline:
            sheet_ranges.cell(column=col, row=1, value=firstline[col-1])
            col = col + 1

    measlist = ['NF', 'R1_1', 'R2_2', 'S21']

    header()
    # instDict['Supply'].__SetEnable__(1)

    instDict['Supply'].__SetEnable__(1)
    time.sleep(0.1)
    pluto = PlutoV0()
    pluto.connect(DUT1_Default=0x00, DUT2_Default=0x00)
    pluto.Set_Amp_Gain(SPI_sel=channel, GainValue='20dB')
    pluto.Set_Amp_Coupling(SPI_sel=channel, Coupling='ON')
    pluto.Set_Amp_Pwr_Mod(SPI_sel=channel, PowerMode="Hi")
    pluto.Set_Amp_Enable(SPI_sel=channel, AmpEnable=True)
    pluto.Set_Amp_Trim(SPI_sel=channel, AmpTrimCode=15)

    for temp in templist:
        if templist != [25]:
            setTemp(temp)
        fh.write('Temp = %d' % temp)
        fh.write('\n')
        for supply in supplylist:
            instDict['Supply'].__SetV__(supply)
            fh.write('Supply = %g\n' % supply)
            for vcom in vcomlist:
                if vcom != 'N/A':
                    instDict['vcom'].__SetV__(vcom)
                    fh.write('Vcom = %g\n' % vcom)
                    time.sleep(0.2)
                for pwrmode in pwrmodelist:
                    fh.write('PowerMode = %s' % pwrmode)
                    pluto.Set_Amp_Pwr_Mod(SPI_sel=channel, PowerMode=pwrmode)
                    for gain in gainlist:
                        fh.write('Gain = %s\n' % gain)
                        pluto.Set_Amp_Gain(SPI_sel=channel, GainValue=gain)
                        VNAinit()
                        getData()

    if templist != [25]:
        setTemp(25)
        instDict['thermo'].__FlowOFF__()

    pluto.Set_Amp_Enable(SPI_sel=channel, AmpEnable=False)
    instDict['Supply'].__SetEnable__(0)
    instDict['NA'].__Output__(0)
    endTime = time.time() - startTime

    print 'Program executed in %d seconds.' % endTime

    fh.write('Execution Time = ,%d' % endTime)

    fh.close()
    wb.save(filename=summaryPath)

if __name__ == '__main__':
    path = 'C:\\Users\\bsulliv2\\Desktop\\Results\\PlutoV0\\NF\\'
    summaryPath = 'C:\\Users\\bsulliv2\\Desktop\\Results\\PlutoV0\\PlutoNFSummary.xlsx'

    avg = 3
    SourceAtten = 10
    P1 = P2 = -30
    noiseBW = 4e6


    numPoints = 201
    startFreq = 10e6
    endFreq = 10.01e9

    templist = [25]
    vcomlist = ['N/A']
    supplylist = [3.3]
    gainlist = ['12dB', '20dB']
    # gainlist = ['20dB']
    pwrmodelist = ['Lo', 'Hi']
    dut = 'VB01 A'
    channel = 'A'

    instDict = InstInit()

    NF()
